import datetime
import os
import sys
from datetime import timedelta

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from PySide6.QtCore import Qt, QDate, QTime, QTimer, QPoint
from PySide6.QtGui import QFont, QPixmap, QIcon, QColor, QTextCharFormat, QGuiApplication
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QLabel,
    QFrame,
    QScrollArea,
    QSizePolicy,
    QPushButton,
    QDateEdit,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QStackedWidget,
    QCalendarWidget,
    QToolButton,
    QMenu,
    QSpinBox,
    QAbstractItemView,
    QFileDialog,
    QMessageBox,
    QDialog,
    QComboBox,
)

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.styles import Alignment as XLAlignment, Border as XLBorder, Font as XLFont, PatternFill as XLPatternFill, Side as XLSide
from openpyxl.utils import get_column_letter

from views.admin_dashboard_view import (
    AMBER,
    AMBER_BG,
    BASE_DIR,
    GRAY_1,
    GRAY_2,
    GRAY_3,
    GRAY_4,
    GRAY_5,
    GREEN,
    GREEN_BG,
    RED,
    RED_BG,
    TEAL,
    TEAL_DARK,
    TEAL_LIGHT,
    TEAL_MID,
    TEAL_PALE,
    WHITE,
    inter,
    playfair,
    owner_scrollbar_qss,
    load_fonts,
)

_FONTS_READY = False


def _ensure_fonts_loaded():
    global _FONTS_READY
    if _FONTS_READY:
        return
    # QFontDatabase access inside load_fonts requires a running Qt application.
    if QGuiApplication.instance() is None:
        return
    load_fonts()
    _FONTS_READY = True

MODERN_CHEVRON_ICON = os.path.join(BASE_DIR, "assets", "chevron_down_modern.svg")
WHITE_CHEVRON_ICON = os.path.join(BASE_DIR, "assets", "chevron_down_white.svg")
CALENDAR_ICON = os.path.join(BASE_DIR, "assets", "calendar_icon.svg")

PDF_FONT_REGULAR = "Inter-Regular"
PDF_FONT_MEDIUM = "Inter-Medium"
PDF_FONT_BOLD = "Inter-SemiBold"
PDF_FONT_PLAYFAIR = "PlayfairDisplay-SemiBold"
_PDF_FONTS_READY = False


def _register_pdf_fonts():
    global _PDF_FONTS_READY
    if _PDF_FONTS_READY:
        return

    font_dir = os.path.join(BASE_DIR, "assets", "fonts")
    font_map = {
        PDF_FONT_REGULAR: os.path.join(font_dir, "Inter_18pt-Regular.ttf"),
        PDF_FONT_MEDIUM: os.path.join(font_dir, "Inter_18pt-Medium.ttf"),
        PDF_FONT_BOLD: os.path.join(font_dir, "Inter_18pt-SemiBold.ttf"),
        PDF_FONT_PLAYFAIR: os.path.join(font_dir, "PlayfairDisplay-SemiBold.ttf"),
    }

    for font_name, font_path in font_map.items():
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont(font_name, font_path))

    _PDF_FONTS_READY = True


def _pdf_font(kind="regular"):
    mapping = {
        "regular": PDF_FONT_REGULAR,
        "medium": PDF_FONT_MEDIUM,
        "bold": PDF_FONT_BOLD,
        "playfair": PDF_FONT_PLAYFAIR,
    }
    requested = mapping.get(kind, PDF_FONT_REGULAR)
    available = set(pdfmetrics.getRegisteredFontNames())
    if requested in available:
        return requested
    if kind == "bold":
        return "Helvetica-Bold"
    return "Helvetica"


def _qss_path(path):
    return path.replace("\\", "/")


def _date_edit_style(min_height=34, min_width=None):
    min_width_rule = f"min-width:{min_width}px;" if min_width else ""
    icon_path = _qss_path(CALENDAR_ICON)
    return f"""
    QDateEdit {{
        color:{GRAY_5};
        background:#fbfcfc;
        border:1px solid #d6e2df;
        border-radius:8px;
        padding:0 40px 0 12px;
        min-height:{min_height}px;
        {min_width_rule}
    }}
    QDateEdit:hover {{
        border-color:#b9d4cf;
        background:{WHITE};
    }}
    QDateEdit:focus {{
        border:1px solid {TEAL};
        background:{WHITE};
    }}
    QDateEdit::drop-down {{
        subcontrol-origin: padding;
        subcontrol-position: top right;
        width:24px;
        right:4px;
        border:none;
        background:transparent;
    }}
    QDateEdit::down-arrow {{
        image:url({icon_path});
        width:14px;
        height:14px;
    }}
    """


def _calendar_native_tint_style():
    inter_family = inter(11).family()
    white_chevron_path = _qss_path(WHITE_CHEVRON_ICON)
    return f"""
    QCalendarWidget {{
        background:{WHITE};
        color:{GRAY_5};
        border:1px solid {GRAY_2};
        font-family:'{inter_family}';
    }}
    QCalendarWidget QWidget#qt_calendar_navigationbar {{
        background:{TEAL_DARK};
        color:{WHITE};
        border-top-left-radius:8px;
        border-top-right-radius:8px;
    }}
    QCalendarWidget QToolButton {{
        color:{WHITE};
        background:transparent;
        border:none;
        padding:4px 8px;
        border-radius:6px;
        font-family:'{inter_family}';
        font-size:12px;
        font-weight:500;
    }}
    QCalendarWidget QToolButton#qt_calendar_monthbutton,
    QCalendarWidget QToolButton#qt_calendar_yearbutton {{
        background:transparent;
        border:none;
        padding:4px 22px 4px 6px;
        border-radius:0px;
        font-family:'{inter_family}';
        font-size:14px;
        font-weight:600;
    }}
    QCalendarWidget QToolButton#qt_calendar_monthbutton::menu-indicator,
    QCalendarWidget QToolButton#qt_calendar_yearbutton::menu-indicator {{
        image:url({white_chevron_path});
        width:10px;
        height:10px;
        subcontrol-origin: padding;
        subcontrol-position: right center;
        right:6px;
    }}
    QCalendarWidget QToolButton:hover {{
        background:rgba(255,255,255,0.16);
    }}
    QCalendarWidget QToolButton:pressed {{
        background:rgba(255,255,255,0.24);
    }}
    QCalendarWidget QSpinBox {{
        color:{WHITE};
        background:rgba(255,255,255,0.12);
        border:1px solid rgba(255,255,255,0.24);
        border-radius:6px;
        padding-right:20px;
        min-height:24px;
        min-width:72px;
        font-family:'{inter_family}';
        font-size:12px;
    }}
    QCalendarWidget QSpinBox::up-button {{
        subcontrol-origin:border;
        subcontrol-position:top right;
        width:18px;
        border-left:1px solid rgba(255,255,255,0.28);
        border-top-right-radius:6px;
        background:rgba(255,255,255,0.10);
    }}
    QCalendarWidget QSpinBox::down-button {{
        subcontrol-origin:border;
        subcontrol-position:bottom right;
        width:18px;
        border-left:1px solid rgba(255,255,255,0.28);
        border-top:1px solid rgba(255,255,255,0.22);
        border-bottom-right-radius:6px;
        background:rgba(255,255,255,0.10);
    }}
    QCalendarWidget QSpinBox::up-button:hover,
    QCalendarWidget QSpinBox::down-button:hover {{
        background:rgba(255,255,255,0.18);
    }}
    QCalendarWidget QWidget#qt_calendar_calendarview {{
        outline:none;
        border:none;
        background:{WHITE};
    }}
    QCalendarWidget QHeaderView::section {{
        background:{WHITE};
        color:{TEAL};
        border:none;
        padding:6px 0;
        font-family:'{inter_family}';
        font-size:10px;
        font-weight:600;
    }}
    QCalendarWidget QAbstractItemView {{
        background:{WHITE};
        color:#000000;
        selection-background-color:{TEAL_PALE};
        selection-color:{TEAL_DARK};
        outline:none;
        font-family:'{inter_family}';
        font-size:12px;
        font-weight:600;
    }}
    QCalendarWidget QAbstractItemView:item {{
        background:transparent;
        border:none;
        padding:2px;
        margin:0px;
    }}
    QCalendarWidget QAbstractItemView:item:hover {{
        background:rgba(26,122,110,0.06);
        border:none;
        color:#000000;
    }}
    QCalendarWidget QAbstractItemView:item:selected {{
        background:{TEAL_PALE};
        border:1px solid {TEAL};
        border-radius:6px;
        color:#000000;
        font-weight:600;
    }}
    """


def _calendar_popup_menu_style():
    inter_family = inter(11).family()
    return f"""
    QMenu {{
        background:{WHITE};
        color:{TEAL_DARK};
        border:1px solid {GRAY_2};
        padding:4px;
        font-family:'{inter_family}';
        font-size:12px;
    }}
    QMenu::item {{
        background:transparent;
        color:{TEAL_DARK};
        padding:6px 14px;
        border-radius:4px;
    }}
    QMenu::item:selected {{
        background:{TEAL_PALE};
        color:{TEAL_DARK};
    }}
    """


def _attach_year_dropdown(calendar):
    year_btn = calendar.findChild(QToolButton, "qt_calendar_yearbutton")
    if year_btn is None:
        return

    year_spin = calendar.findChild(QSpinBox, "qt_calendar_yearedit")
    if year_spin is not None:
        year_spin.hide()

    menu = QMenu(year_btn)
    menu.setStyleSheet(_calendar_popup_menu_style())

    def refill(target_year):
        menu.clear()
        for year in range(target_year - 15, target_year + 16):
            action = menu.addAction(str(year))
            action.setData(year)

    def on_year_picked(action):
        target_year = int(action.data())
        current_month = calendar.monthShown()
        calendar.setCurrentPage(target_year, current_month)

        selected = calendar.selectedDate()
        if selected.isValid():
            target_day = min(selected.day(), QDate(target_year, current_month, 1).daysInMonth())
            calendar.setSelectedDate(QDate(target_year, current_month, target_day))
        year_btn.setText(str(target_year))
        refill(target_year)

    def on_page_changed(year, month):
        year_btn.setText(str(year))
        refill(year)

    refill(calendar.yearShown())
    year_btn.setText(str(calendar.yearShown()))
    year_btn.setPopupMode(QToolButton.InstantPopup)
    year_btn.setMenu(menu)
    menu.triggered.connect(on_year_picked)
    calendar.currentPageChanged.connect(on_page_changed)


class NumberedCanvas(pdf_canvas.Canvas):
    def __init__(self, *args, footer_text="", footer_font_name="Helvetica", **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
        self._footer_text = footer_text
        self._footer_font_name = footer_font_name

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for page_state in self._saved_page_states:
            self.__dict__.update(page_state)
            self._draw_page_footer(page_count)
            super().showPage()
        super().save()

    def _draw_page_footer(self, page_count):
        width, height = landscape(letter)
        self.setStrokeColor(colors.HexColor("#e6eae9"))
        self.setLineWidth(0.6)
        self.line(30, 35, width - 30, 35)

        self.setFont(self._footer_font_name, 8.5)
        self.setFillColor(colors.HexColor("#3a4a47"))
        self.drawString(30, 22, self._footer_text)
        self.drawRightString(width - 30, 22, f"Page {self._pageNumber} of {page_count}")


class ExportNoticeDialog(QDialog):
    def __init__(self, title_text, message, parent=None):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowTitle("")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint | Qt.CustomizeWindowHint)
        self.setStyleSheet("background:transparent;")
        self.setMinimumWidth(430)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        card = QFrame()
        card.setStyleSheet(
            f"""
            QFrame {{
                background:{WHITE};
                border:1px solid {GRAY_2};
                border-radius:10px;
            }}
            """
        )
        root.addWidget(card)

        lay = QVBoxLayout(card)
        lay.setContentsMargins(22, 20, 22, 18)
        lay.setSpacing(12)

        title = QLabel(title_text)
        title.setFont(playfair(16, QFont.Medium))
        title.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")

        body = QLabel(str(message))
        body.setWordWrap(True)
        body.setFont(inter(11))
        body.setStyleSheet(f"color:{GRAY_5};background:transparent;border:none;")

        btn_row = QHBoxLayout()
        btn_row.addStretch()

        ok_btn = QPushButton("OK")
        ok_btn.setCursor(Qt.PointingHandCursor)
        ok_btn.setFont(inter(10, QFont.Medium))
        ok_btn.setFixedSize(84, 34)
        ok_btn.setStyleSheet(
            f"""
            QPushButton {{
                background:{TEAL};
                color:{WHITE};
                border:1px solid {TEAL};
                border-radius:6px;
            }}
            QPushButton:hover {{
                background:{TEAL_MID};
            }}
            """
        )
        ok_btn.clicked.connect(self.accept)
        btn_row.addWidget(ok_btn)

        lay.addWidget(title)
        lay.addWidget(body)
        lay.addLayout(btn_row)


class StatusBadge(QLabel):
    """Status badge widget for delivery status"""
    def __init__(self, status, parent=None):
        super().__init__(parent)
        self.setText(status)
        self.setAlignment(Qt.AlignCenter)
        self.setFont(inter(10,QFont.Medium))
        self.setFixedHeight(26)
        self.setFixedWidth(94)

        status_lower = status.lower()
        if status_lower == "delivered":
            self.setStyleSheet(
                f"background:{GREEN_BG};color:{GREEN};border:1px solid {GREEN};border-radius:4px;"
            )
        elif status_lower == "in transit":
            self.setStyleSheet(
                f"background:{TEAL_PALE};color:{TEAL_DARK};border:1px solid {TEAL};border-radius:4px;"
            )
        elif status_lower == "pending":
            self.setStyleSheet(
                f"background:{AMBER_BG};color:{AMBER};border:1px solid {AMBER};border-radius:4px;"
            )
        elif status_lower == "cancelled":
            self.setStyleSheet(
                f"background:{RED_BG};color:{RED};border:1px solid {RED};border-radius:4px;"
            )
        else:
            self.setStyleSheet(
                f"background:{GRAY_2};color:{GRAY_4};border:1px solid {GRAY_3};border-radius:4px;"
            )


class PaymentBadge(QLabel):
    """Payment status badge for report rows."""
    def __init__(self, status, parent=None):
        super().__init__(parent)
        status = status or "Not Applicable"
        self.setText(status)
        self.setAlignment(Qt.AlignCenter)
        self.setFont(inter(11, QFont.Medium))
        self.setFixedHeight(26)
        self.setFixedWidth(156)

        status_lower = status.lower()
        if status_lower == "paid":
            self.setStyleSheet(
                f"background:{GREEN_BG};color:{GREEN};border:1px solid {GREEN};border-radius:4px;"
            )
        elif status_lower == "unpaid":
            self.setStyleSheet(
                f"background:{AMBER_BG};color:{AMBER};border:1px solid {AMBER};border-radius:4px;"
            )
        else:
            self.setStyleSheet(
                f"background:{GRAY_2};color:{GRAY_4};border:1px solid {GRAY_3};border-radius:4px;"
            )


class ReportTabPanel(QFrame):
    """Tab panel for each report period (Daily, Weekly, Monthly)"""
    def __init__(self, period_name, start_date, end_date, owner_name="Owner", parent=None):
        super().__init__(parent)
        _ensure_fonts_loaded()
        self.period_name = period_name
        self.start_date = start_date
        self.end_date = end_date
        self.owner_name = owner_name or "Owner"
        self._current_from_date = start_date
        self._current_to_date = end_date
        self._summary_data = {
            "total_deliveries": 0,
            "total_delivered": 0,
            "total_cancelled": 0,
            "total_pending": 0,
            "total_in_transit": 0,
            "total_sales": 0.0,
            "total_paid": 0.0,
            "total_unpaid": 0.0,
            "avg_transaction_value": 0.0,
            "peak_sales_day": "",
            "peak_sales_amount": 0.0,
            "most_sold_product": "",
            "most_sold_product_quantity": 0,
            "most_sold_product_revenue": 0.0,
        }
        self._base_data = []
        self._filtered_data = []

        self.setStyleSheet("background:transparent;border:none;")
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(16)

        # Summary blocks
        sales_frame = QFrame()
        sales_frame.setStyleSheet(
            f"""
            QFrame{{
                background:{WHITE};
                border:1px solid {GRAY_2};
                border-radius:10px;
            }}
            """
        )
        sales_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sales_frame.setFixedHeight(224)

        sales_root = QVBoxLayout(sales_frame)
        sales_root.setContentsMargins(0, 0, 0, 0)
        sales_root.setSpacing(0)

        sales_top = QFrame()
        sales_top.setFixedHeight(4)
        sales_top.setStyleSheet(
            f"background:{TEAL};border:none;border-top-left-radius:10px;border-top-right-radius:10px;"
        )
        sales_root.addWidget(sales_top)

        sales_body = QWidget()
        sales_body.setStyleSheet("background:transparent;border:none;")
        sales_lay = QVBoxLayout(sales_body)
        sales_lay.setContentsMargins(24, 17, 24, 18)
        sales_lay.setSpacing(7)

        sales_title_text, sales_desc_text = self._get_sales_texts()

        sales_lbl = QLabel(sales_title_text)
        sales_lbl.setFont(inter(11, QFont.DemiBold))
        sales_lbl.setStyleSheet(
            f"color:{GRAY_4};letter-spacing:1px;background:transparent;border:none;"
        )
        self._sales_title_lbl = sales_lbl
        total_sales = 0

        sales_val = QLabel(f"PHP {total_sales:,.2f}")
        sales_val.setFont(playfair(38, QFont.Medium))
        sales_val.setMinimumHeight(48)
        sales_val.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")
        self._sales_val_lbl = sales_val

        sales_desc = QLabel(sales_desc_text)
        sales_desc.setFont(inter(11))
        sales_desc.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")
        self._sales_desc_lbl = sales_desc

        sales_lay.addWidget(sales_lbl)
        sales_lay.addWidget(sales_val)
        sales_lay.addWidget(sales_desc)
        money_row = QHBoxLayout()
        money_row.setContentsMargins(0, 10, 0, 0)
        money_row.setSpacing(26)
        collected_item, self._paid_value_lbl = self._summary_value("Collected", "PHP 0.00", GREEN)
        receivable_item, self._unpaid_value_lbl = self._summary_value("Receivables", "PHP 0.00", AMBER)
        money_row.addWidget(collected_item, 1)
        money_row.addWidget(receivable_item, 1)
        sales_lay.addLayout(money_row)
        sales_lay.addStretch()
        sales_root.addWidget(sales_body)
        delivery_frame, delivery_lay = self._summary_panel("DELIVERY SUMMARY", TEAL_MID, height=224)
        delivery_lay.setSpacing(5)
        self._delivery_total_lbl = QLabel("0")
        self._delivery_total_lbl.setFont(inter(28, QFont.DemiBold))
        self._delivery_total_lbl.setFixedHeight(44)
        self._delivery_total_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self._delivery_total_lbl.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")
        delivery_lay.addWidget(self._delivery_total_lbl)
        total_hint = QLabel("Total deliveries in this report")
        total_hint.setFont(inter(10))
        total_hint.setFixedHeight(16)
        total_hint.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        total_hint.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")
        delivery_lay.addWidget(total_hint)

        delivery_grid = QGridLayout()
        delivery_grid.setContentsMargins(0, 8, 0, 0)
        delivery_grid.setHorizontalSpacing(18)
        delivery_grid.setVerticalSpacing(6)
        delivered_item, self._delivered_value_lbl = self._summary_value("Delivered", "0", GREEN)
        pending_item, self._pending_value_lbl = self._summary_value("Pending", "0", AMBER)
        transit_item, self._in_transit_value_lbl = self._summary_value("In transit", "0", TEAL_MID)
        cancelled_item, self._cancelled_value_lbl = self._summary_value("Cancelled", "0", RED)
        delivery_grid.addWidget(delivered_item, 0, 0)
        delivery_grid.addWidget(pending_item, 0, 1)
        delivery_grid.addWidget(transit_item, 1, 0)
        delivery_grid.addWidget(cancelled_item, 1, 1)
        delivery_lay.addLayout(delivery_grid)
        delivery_lay.addStretch()

        insights_frame, insights_lay = self._summary_panel("REPORT INSIGHTS", AMBER, height=224)
        peak_item, self._peak_day_value_lbl, self._peak_day_detail_lbl = self._insight_row("Peak sales day")
        product_item, self._most_product_value_lbl, self._most_product_detail_lbl = self._insight_row("Most sold product")
        avg_item, self._avg_order_value_lbl, self._avg_order_detail_lbl = self._insight_row("Average order value")
        insights_lay.addWidget(peak_item)
        insights_lay.addWidget(product_item)
        insights_lay.addWidget(avg_item)
        insights_lay.addStretch()

        summary_grid = QGridLayout()
        summary_grid.setContentsMargins(0, 0, 0, 0)
        summary_grid.setHorizontalSpacing(16)
        summary_grid.setVerticalSpacing(16)
        summary_grid.addWidget(sales_frame, 0, 0)
        summary_grid.addWidget(delivery_frame, 0, 1)
        summary_grid.addWidget(insights_frame, 0, 2)
        summary_grid.setColumnStretch(0, 2)
        summary_grid.setColumnStretch(1, 1)
        summary_grid.setColumnStretch(2, 1)
        lay.addLayout(summary_grid)

        # Table Section Header
        table_header = QHBoxLayout()
        table_header.setContentsMargins(0, 8, 0, 8)
        table_header.setSpacing(8)

        table_title = QLabel("Delivery Details")
        table_title.setFont(playfair(14, QFont.Medium))
        table_title.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")
        table_header.addWidget(table_title)
        table_header.addStretch()

        self._status_filter = self._build_filter_combo(
            [
                ("All Statuses", "All"),
                ("Delivered", "Delivered"),
                ("Pending", "Pending"),
                ("In Transit", "In Transit"),
                ("Cancelled", "Cancelled"),
            ],
            132,
        )
        self._status_filter.currentIndexChanged.connect(self._apply_local_filters)
        table_header.addWidget(self._status_filter)

        self._payment_filter = self._build_filter_combo(
            [
                ("All Payments", "All"),
                ("Paid", "Paid"),
                ("Unpaid", "Unpaid"),
                ("Not Applicable", "Not Applicable"),
            ],
            146,
        )
        self._payment_filter.currentIndexChanged.connect(self._apply_local_filters)
        table_header.addWidget(self._payment_filter)

        export_pdf_btn = QPushButton("Export PDF")
        export_pdf_btn.setFont(inter(9, QFont.Medium))
        export_pdf_btn.setFixedSize(92, 32)
        export_pdf_btn.setCursor(Qt.PointingHandCursor)
        export_pdf_btn.setStyleSheet(
            f"""
            QPushButton{{
                background:{TEAL};color:#fff;border:none;border-radius:4px;
            }}
            QPushButton:hover{{background:{TEAL_MID};}}
        """
        )
        export_pdf_btn.clicked.connect(self._export_pdf)
        table_header.addWidget(export_pdf_btn)

        export_excel_btn = QPushButton("Export Excel")
        export_excel_btn.setFont(inter(9, QFont.Medium))
        export_excel_btn.setFixedSize(100, 32)
        export_excel_btn.setCursor(Qt.PointingHandCursor)
        export_excel_btn.setStyleSheet(
            f"""
            QPushButton{{
                background:{WHITE};color:{TEAL_DARK};border:1px solid {TEAL_LIGHT};border-radius:4px;
            }}
            QPushButton:hover{{background:{TEAL_PALE};}}
        """
        )
        export_excel_btn.clicked.connect(self._export_excel)
        table_header.addWidget(export_excel_btn)

        lay.addLayout(table_header)

        # Table
        self.table = QTableWidget()
        self.table.setObjectName("reportDeliveryTable")
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["Delivery ID", "Customer", "Product", "Status", "Payment", "Amount", "Date"]
        )
        for col_idx in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col_idx)
            if header_item is None:
                continue
            if col_idx in (3, 4):
                header_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            elif col_idx == 5:
                header_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            else:
                header_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setFont(inter(11))
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(56)
        self.table.verticalHeader().setMinimumSectionSize(52)
        self.table.setWordWrap(False)
        self.table.setShowGrid(False)
        self.table.setAlternatingRowColors(True)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        header = self.table.horizontalHeader()
        header.setFont(inter(11, QFont.DemiBold))
        header.setFixedHeight(48)
        header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        header.setHighlightSections(False)
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        header.setSectionResizeMode(4, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        header.setSectionResizeMode(6, QHeaderView.Fixed)
        self.table.setStyleSheet(
            f"""
            QTableWidget#reportDeliveryTable{{
                background:{WHITE};
                alternate-background-color:#f8fbfa;
                border:1px solid {GRAY_2};
                border-radius:8px;
                gridline-color:transparent;
            }}
            QTableWidget#reportDeliveryTable QHeaderView::section{{
                background:#f4f8f7;
                color:{GRAY_5};
                padding:0 12px;
                font-size:14px;
                font-weight:600;
                border:none;
                border-bottom:1px solid {GRAY_2};
            }}
            QTableWidget#reportDeliveryTable::item{{
                color:{GRAY_5};
                padding:0 12px;
                font-size:14px;
                border:none;
                border-bottom:1px solid #edf2f1;
            }}
            QTableWidget#reportDeliveryTable::item:selected{{
                background:{TEAL_PALE};color:{TEAL_DARK};
            }}
            QTableWidget#reportDeliveryTable::item:hover{{
                background:#eef8f6;
            }}
            QTableWidget#reportDeliveryTable::item:focus{{
                outline:none;
                border:none;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: 8px;
                margin: 2px 2px 2px 0;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {TEAL};
                min-height: 24px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {TEAL_MID};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
                background: transparent;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: transparent;
            }}
            QScrollBar:horizontal {{
                background: transparent;
                height: 8px;
                margin: 0 2px 2px 2px;
                border: none;
            }}
            QScrollBar::handle:horizontal {{
                background: {TEAL};
                min-width: 24px;
                border-radius: 4px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {TEAL_MID};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                border: none;
                background: transparent;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: transparent;
            }}
        """
        )
        self.table.setStyleSheet(self.table.styleSheet() + owner_scrollbar_qss())
        self.table.setFixedHeight(488)
        self.table.setColumnWidth(0, 104)
        self.table.setColumnWidth(3, 132)
        self.table.setColumnWidth(4, 188)
        self.table.setColumnWidth(5, 148)
        self.table.setColumnWidth(6, 138)

        self._populate_table()
        lay.addWidget(self.table)

    def _summary_panel(self, title, top_color, height=224):
        frame = QFrame()
        frame.setStyleSheet(
            f"""
            QFrame{{
                background:{WHITE};
                border:1px solid {GRAY_2};
                border-radius:8px;
            }}
            """
        )
        frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        frame.setFixedHeight(height)

        root = QVBoxLayout(frame)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        top = QFrame()
        top.setFixedHeight(4)
        top.setStyleSheet(
            f"background:{top_color};border:none;border-top-left-radius:8px;border-top-right-radius:8px;"
        )
        root.addWidget(top)

        body = QWidget()
        body.setStyleSheet("background:transparent;border:none;")
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(20, 15, 20, 18)
        body_lay.setSpacing(8)

        title_lbl = QLabel(title)
        title_lbl.setFont(inter(10, QFont.DemiBold))
        title_lbl.setStyleSheet(f"color:{GRAY_4};letter-spacing:1px;background:transparent;border:none;")
        body_lay.addWidget(title_lbl)

        root.addWidget(body)
        return frame, body_lay

    def _summary_value(self, label, value, color=TEAL_DARK):
        item = QWidget()
        item.setStyleSheet("background:transparent;border:none;")
        item.setMinimumHeight(38)
        lay = QVBoxLayout(item)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(1)

        label_lbl = QLabel(label)
        label_lbl.setFont(inter(10, QFont.Medium))
        label_lbl.setMinimumHeight(15)
        label_lbl.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")

        value_lbl = QLabel(value)
        value_lbl.setFont(inter(14, QFont.DemiBold))
        value_lbl.setMinimumHeight(21)
        value_lbl.setStyleSheet(f"color:{color};background:transparent;border:none;")

        lay.addWidget(label_lbl)
        lay.addWidget(value_lbl)
        return item, value_lbl

    def _insight_row(self, label):
        item = QWidget()
        item.setStyleSheet("background:transparent;border:none;")
        item.setMinimumHeight(48)
        item.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        lay = QHBoxLayout(item)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(12)

        label_lbl = QLabel(label)
        label_lbl.setFont(inter(10, QFont.Medium))
        label_lbl.setFixedWidth(132)
        label_lbl.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        label_lbl.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")

        value_col = QVBoxLayout()
        value_col.setContentsMargins(0, 0, 0, 0)
        value_col.setSpacing(1)

        value_lbl = QLabel("-")
        value_lbl.setFont(inter(14, QFont.DemiBold))
        value_lbl.setMinimumHeight(22)
        value_lbl.setWordWrap(False)
        value_lbl.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")

        detail_lbl = QLabel("")
        detail_lbl.setFont(inter(9))
        detail_lbl.setMinimumHeight(14)
        detail_lbl.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")

        lay.addWidget(label_lbl)
        value_col.addWidget(value_lbl)
        value_col.addWidget(detail_lbl)
        lay.addLayout(value_col, 1)
        return item, value_lbl, detail_lbl

    def _build_filter_combo(self, items, width):
        combo = QComboBox()
        combo.setFont(inter(9, QFont.Medium))
        combo.setFixedSize(width, 32)
        combo.setCursor(Qt.PointingHandCursor)
        for label, data in items:
            combo.addItem(label, data)
        combo.setStyleSheet(
            f"""
            QComboBox{{
                background:{WHITE};
                color:{TEAL_DARK};
                border:1px solid {TEAL_LIGHT};
                border-radius:4px;
                padding:0 28px 0 10px;
            }}
            QComboBox:hover{{background:{TEAL_PALE};}}
            QComboBox::drop-down{{
                width:24px;
                border:none;
                background:transparent;
            }}
            QComboBox QAbstractItemView{{
                background:{WHITE};
                color:{GRAY_5};
                border:1px solid {GRAY_2};
                selection-background-color:{TEAL_PALE};
                selection-color:{TEAL_DARK};
            }}
            """
        )
        return combo

    def _get_sales_texts(self):
        from_date = self._current_from_date
        to_date = self._current_to_date
        if from_date == to_date:
            range_text = from_date.toString("MMMM d, yyyy")
        else:
            range_text = f"{from_date.toString('MMMM d, yyyy')} - {to_date.toString('MMMM d, yyyy')}"

        if self.period_name == "Daily":
            return "DELIVERED SALES", f"Delivered sales ({range_text})"

        if self.period_name == "Weekly":
            return "DELIVERED SALES THIS WEEK", f"Delivered sales ({range_text})"

        return "DELIVERED SALES THIS MONTH", f"Delivered sales ({range_text})"

    def _refresh_sales_labels(self):
        title_text, desc_text = self._get_sales_texts()
        self._sales_title_lbl.setText(title_text)
        self._sales_desc_lbl.setText(desc_text)

    def _populate_table(self):
        """Populate table with delivery data"""
        self.table.setSortingEnabled(False)
        self.table.clearSpans()
        self.table.setRowCount(0)
        if not self._filtered_data:
            self.table.setRowCount(1)
            self.table.setSpan(0, 0, 1, self.table.columnCount())
            empty_item = QTableWidgetItem("No report records found for the selected filters.")
            empty_item.setTextAlignment(Qt.AlignCenter)
            empty_item.setFont(inter(10, QFont.Medium))
            empty_item.setForeground(QColor(GRAY_4))
            self.table.setItem(0, 0, empty_item)
            return

        for row_idx, delivery in enumerate(self._filtered_data):
            self.table.insertRow(row_idx)
            self.table.setRowHeight(row_idx, 56)

            # Delivery ID
            id_item = QTableWidgetItem(str(delivery.get("id", "")))
            id_item.setFont(inter(11, QFont.Medium))
            id_item.setForeground(QColor(TEAL_DARK))
            id_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row_idx, 0, id_item)

            # Customer
            cust_item = QTableWidgetItem(str(delivery.get("customer", "")))
            cust_item.setFont(inter(11))
            cust_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row_idx, 1, cust_item)

            # Product
            prod_item = QTableWidgetItem(str(delivery.get("product", "")))
            prod_item.setFont(inter(11))
            prod_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row_idx, 2, prod_item)

            # Status
            status = str(delivery.get("status", ""))
            status_item = QTableWidgetItem("")
            status_item.setData(Qt.UserRole, status)
            status_item.setFont(inter(11))
            self.table.setItem(row_idx, 3, status_item)
            status_widget = QWidget()
            status_widget.setStyleSheet("background:transparent;border:none;")
            status_layout = QHBoxLayout(status_widget)
            status_layout.setContentsMargins(0, 0, 0, 0)
            status_layout.setAlignment(Qt.AlignCenter)
            status_badge = StatusBadge(status)
            status_layout.addWidget(status_badge)
            self.table.setCellWidget(row_idx, 3, status_widget)

            # Payment
            payment_status = self._payment_status_for_row(delivery)
            payment_item = QTableWidgetItem("")
            payment_item.setData(Qt.UserRole, payment_status)
            payment_item.setFont(inter(11))
            self.table.setItem(row_idx, 4, payment_item)
            payment_widget = QWidget()
            payment_widget.setStyleSheet("background:transparent;border:none;")
            payment_layout = QHBoxLayout(payment_widget)
            payment_layout.setContentsMargins(0, 0, 0, 0)
            payment_layout.setAlignment(Qt.AlignCenter)
            payment_layout.addWidget(PaymentBadge(payment_status))
            self.table.setCellWidget(row_idx, 4, payment_widget)

            # Amount
            amount = float(delivery.get("amount", 0.0) or 0.0)
            amount_item = QTableWidgetItem(f"PHP {amount:,.2f}")
            amount_item.setFont(inter(11, QFont.Medium))
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            amount_item.setForeground(QColor(GREEN if status == "Delivered" else GRAY_4))
            self.table.setItem(row_idx, 5, amount_item)

            # Date
            date_item = QTableWidgetItem(self._format_table_date(delivery.get("date", "")))
            date_item.setFont(inter(11))
            date_item.setForeground(QColor(GRAY_5))
            date_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row_idx, 6, date_item)

        self.table.setSortingEnabled(True)

    def _format_table_date(self, raw_date):
        raw_text = str(raw_date or "").strip()
        parsed = QDate.fromString(raw_text, "yyyy-MM-dd")
        if parsed.isValid():
            return parsed.toString("MMM d, yyyy")
        return raw_text

    def apply_date_filter(self, from_date, to_date):
        """Filter this tab's data by selected date range and refresh KPIs/table."""
        self._current_from_date = from_date
        self._current_to_date = to_date
        self._apply_local_filters()

    def _selected_filter_data(self, combo):
        if combo is None:
            return "All"
        value = combo.currentData()
        return value if value is not None else "All"

    def _row_date(self, row):
        raw_date = str(row.get("date", "")).strip()
        row_date = QDate.fromString(raw_date, "yyyy-MM-dd")
        if row_date.isValid():
            return row_date
        row_date = QDate.fromString(raw_date, "MMM d, yyyy")
        return row_date if row_date.isValid() else QDate()

    def _payment_status_for_row(self, row):
        payment_status = str(row.get("payment_status") or "").strip()
        if payment_status:
            return payment_status
        return "Unpaid" if row.get("status") == "Delivered" else "Not Applicable"

    def _compact_label(self, value, limit=26):
        text = str(value or "").strip()
        if not text:
            return "-"
        if len(text) <= limit:
            return text
        return f"{text[:max(0, limit - 3)]}..."

    def _is_unfiltered(self):
        return (
            self._selected_filter_data(getattr(self, "_status_filter", None)) == "All"
            and self._selected_filter_data(getattr(self, "_payment_filter", None)) == "All"
        )

    def _apply_local_filters(self, *args, prefer_summary=False):
        status_filter = self._selected_filter_data(getattr(self, "_status_filter", None))
        payment_filter = self._selected_filter_data(getattr(self, "_payment_filter", None))

        filtered_rows = []
        for row in self._base_data:
            row_date = self._row_date(row)
            if not row_date.isValid():
                continue
            if not (self._current_from_date <= row_date <= self._current_to_date):
                continue
            if status_filter != "All" and row.get("status") != status_filter:
                continue
            if payment_filter != "All" and self._payment_status_for_row(row) != payment_filter:
                continue
            filtered_rows.append(row)

        self._filtered_data = filtered_rows
        if prefer_summary and self._is_unfiltered():
            self._refresh_metrics(self._summary_data)
        else:
            self._refresh_metrics(self._metrics_from_rows(filtered_rows))
        self._populate_table()

    def _metrics_from_rows(self, rows):
        statuses_by_delivery = {}
        for index, row in enumerate(rows):
            key = str(row.get("id") or f"row-{index}")
            statuses_by_delivery.setdefault(key, row.get("status", ""))

        total_delivered = sum(1 for status in statuses_by_delivery.values() if status == "Delivered")
        total_pending = sum(1 for status in statuses_by_delivery.values() if status == "Pending")
        total_in_transit = sum(1 for status in statuses_by_delivery.values() if status == "In Transit")
        total_cancelled = sum(1 for status in statuses_by_delivery.values() if status == "Cancelled")

        delivered_rows = [row for row in rows if row.get("status") == "Delivered"]
        total_sales = sum(float(row.get("amount", 0.0) or 0.0) for row in delivered_rows)
        total_paid = sum(
            float(row.get("amount", 0.0) or 0.0)
            for row in delivered_rows
            if self._payment_status_for_row(row) == "Paid"
        )
        total_unpaid = sum(
            float(row.get("amount", 0.0) or 0.0)
            for row in delivered_rows
            if self._payment_status_for_row(row) == "Unpaid"
        )

        delivery_totals = {}
        sales_by_day = {}
        product_totals = {}
        for row in delivered_rows:
            amount = float(row.get("amount", 0.0) or 0.0)
            delivery_id = str(row.get("id") or "")
            if delivery_id:
                delivery_totals[delivery_id] = delivery_totals.get(delivery_id, 0.0) + amount

            row_date = self._row_date(row)
            if row_date.isValid():
                day_name = row_date.toString("dddd")
                sales_by_day[day_name] = sales_by_day.get(day_name, 0.0) + amount

            product_name = str(row.get("product") or "").strip()
            if product_name:
                quantity = int(row.get("quantity") or self._infer_quantity(product_name) or 0)
                product_row = product_totals.setdefault(
                    product_name,
                    {"quantity": 0, "revenue": 0.0},
                )
                product_row["quantity"] += quantity
                product_row["revenue"] += amount

        avg_transaction_value = (
            sum(delivery_totals.values()) / len(delivery_totals)
            if delivery_totals
            else 0.0
        )
        peak_sales_day = ""
        peak_sales_amount = 0.0
        if sales_by_day:
            peak_sales_day, peak_sales_amount = max(
                sales_by_day.items(),
                key=lambda item: (item[1], item[0]),
            )

        most_sold_product = ""
        most_sold_product_quantity = 0
        most_sold_product_revenue = 0.0
        if product_totals:
            most_sold_product, most_sold_values = max(
                product_totals.items(),
                key=lambda item: (item[1]["quantity"], item[1]["revenue"], item[0]),
            )
            most_sold_product_quantity = most_sold_values["quantity"]
            most_sold_product_revenue = most_sold_values["revenue"]

        return {
            "total_deliveries": len(statuses_by_delivery),
            "total_delivered": total_delivered,
            "total_pending": total_pending,
            "total_in_transit": total_in_transit,
            "total_cancelled": total_cancelled,
            "total_sales": total_sales,
            "total_paid": total_paid,
            "total_unpaid": total_unpaid,
            "avg_transaction_value": avg_transaction_value,
            "peak_sales_day": peak_sales_day,
            "peak_sales_amount": peak_sales_amount,
            "most_sold_product": most_sold_product,
            "most_sold_product_quantity": most_sold_product_quantity,
            "most_sold_product_revenue": most_sold_product_revenue,
        }

    def _export_pdf(self):
        default_name = f"{self.period_name.lower()}_report.pdf"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Report as PDF", default_name, "PDF Files (*.pdf)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".pdf"):
            file_path += ".pdf"

        try:
            self._write_pdf(file_path)
            ExportNoticeDialog("Export Successful", f"PDF exported successfully to:\n{file_path}", self).exec()
        except Exception as exc:
            ExportNoticeDialog("Export Failed", f"Could not export PDF report.\n\n{exc}", self).exec()

    def _report_type_label(self):
        return {
            "Daily": "Daily",
            "Weekly": "Weekly",
            "Monthly": "Monthly",
        }.get(self.period_name, self.period_name)

    def _current_range_text(self):
        return f"{self._current_from_date.toString('MMMM d, yyyy')} to {self._current_to_date.toString('MMMM d, yyyy')}"

    def _generated_on_text(self):
        generated_date = QDate.currentDate().toString("MMMM d, yyyy")
        generated_time = QTime.currentTime().toString("h:mm AP")
        return f"Generated on {generated_date} at {generated_time}"

    def _infer_quantity(self, product_name):
        product_lower = product_name.lower()
        if "50kg" in product_lower:
            return 1
        if "22kg" in product_lower:
            return 2
        return 1

    def _report_rows(self):
        rows = []
        for row in self._filtered_data:
            payment_status = self._payment_status_for_row(row)
            delivery_type = row.get("type") or row.get("delivery_type") or ""
            row_date = self._row_date(row)
            date_text = row_date.toString("MMM d, yyyy") if row_date.isValid() else str(row.get("date", ""))
            rows.append(
                {
                    "date": date_text,
                    "customer": row.get("customer", ""),
                    "product": row.get("product", ""),
                    "quantity": row.get("quantity") or self._infer_quantity(row.get("product", "")),
                    "type": delivery_type,
                    "amount": float(row.get("amount", 0.0)),
                    "payment_status": payment_status,
                    "delivery_status": row.get("status", ""),
                }
            )
        return rows

    def _summary_metrics(self):
        return self._metrics_from_rows(self._filtered_data)

    def _insight_values(self, summary):
        peak_day = str(summary.get("peak_sales_day") or "").strip() or "-"
        peak_amount = float(summary.get("peak_sales_amount", 0) or 0)
        product = str(summary.get("most_sold_product") or "").strip() or "-"
        product_quantity = int(summary.get("most_sold_product_quantity", 0) or 0)
        avg_value = float(summary.get("avg_transaction_value", 0) or 0)
        return {
            "peak_day": peak_day,
            "peak_amount": peak_amount,
            "product": product,
            "product_quantity": product_quantity,
            "avg_value": avg_value,
        }

    def _export_metadata(self):
        return {
            "business_name": "G and C LPG Trading",
            "address": "Tuy, Batangas",
            "report_type": self._report_type_label(),
            "date_range": self._current_range_text(),
            "generated_on": self._generated_on_text(),
            "generated_by": self.owner_name,
        }

    def _write_pdf(self, file_path):
        _register_pdf_fonts()
        regular_font = _pdf_font("regular")
        medium_font = _pdf_font("medium")
        bold_font = _pdf_font("bold")
        playfair_font = _pdf_font("playfair")

        metadata = self._export_metadata()
        summary = self._summary_metrics()
        rows = self._report_rows()

        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(letter),
            leftMargin=30,
            rightMargin=30,
            topMargin=28,
            bottomMargin=48,
        )

        styles = getSampleStyleSheet()
        business_style = ParagraphStyle(
            "BusinessName",
            parent=styles["Title"],
            fontName=playfair_font,
            fontSize=18,
            leading=22,
            textColor=colors.HexColor(TEAL_DARK),
            spaceAfter=2,
        )
        address_style = ParagraphStyle(
            "Address",
            parent=styles["Normal"],
            fontName=regular_font,
            fontSize=10,
            leading=12,
            textColor=colors.HexColor("#3a4a47"),
            spaceAfter=6,
        )
        report_type_style = ParagraphStyle(
            "ReportType",
            parent=styles["Heading2"],
            fontName=bold_font,
            fontSize=15,
            leading=18,
            textColor=colors.HexColor(TEAL_DARK),
            spaceAfter=2,
        )
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontName=regular_font,
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor("#263734"),
            spaceAfter=2,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading3"],
            fontName=medium_font,
            fontSize=11.5,
            leading=14,
            textColor=colors.HexColor(TEAL_DARK),
            spaceAfter=6,
        )
        cell_style = ParagraphStyle(
            "Cell",
            parent=styles["Normal"],
            fontName=regular_font,
            fontSize=8.3,
            leading=10,
            textColor=colors.HexColor("#1f2f2c"),
        )
        cell_bold_style = ParagraphStyle(
            "CellBold",
            parent=cell_style,
            fontName=bold_font,
        )
        breakdown_header_style = ParagraphStyle(
            "BreakdownHeader",
            parent=cell_bold_style,
            textColor=colors.white,
        )

        story = [
            Paragraph(metadata["business_name"], business_style),
            Paragraph(metadata["address"], address_style),
            Paragraph(f"{metadata['report_type']} Report", report_type_style),
            Paragraph(f"Date range covered: {metadata['date_range']}", info_style),
            Paragraph(metadata["generated_on"], info_style),
            Paragraph(f"Generated by: {metadata['generated_by']}", info_style),
            Spacer(1, 10),
            Paragraph("Summary", section_style),
        ]

        summary_headers = [
            "Total deliveries",
            "Delivered",
            "Cancelled",
            "Pending",
            "In transit",
            "Delivered sales (PHP)",
            "Total paid (PHP)",
            "Total unpaid (PHP)",
        ]
        summary_values = [
            str(summary["total_deliveries"]),
            str(summary["total_delivered"]),
            str(summary["total_cancelled"]),
            str(summary["total_pending"]),
            str(summary["total_in_transit"]),
            f"{summary['total_sales']:,.2f}",
            f"{summary['total_paid']:,.2f}",
            f"{summary['total_unpaid']:,.2f}",
        ]
        summary_table = Table(
            [
                [Paragraph(text, cell_bold_style) for text in summary_headers],
                [Paragraph(text, cell_style) for text in summary_values],
            ],
            colWidths=[doc.width / 8.0] * 8,
            hAlign="LEFT",
        )
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(TEAL_PALE)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(TEAL_DARK)),
                    ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#1f2f2c")),
                    ("FONTNAME", (0, 0), (-1, -1), regular_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("LEADING", (0, 0), (-1, -1), 10),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(GRAY_2)),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        insights = self._insight_values(summary)
        insight_table = Table(
            [
                [
                    Paragraph("Peak sales day", cell_bold_style),
                    Paragraph("Most sold product", cell_bold_style),
                    Paragraph("Average order value (PHP)", cell_bold_style),
                ],
                [
                    Paragraph(
                        f"{insights['peak_day']}<br/>{insights['peak_amount']:,.2f}",
                        cell_style,
                    ),
                    Paragraph(
                        f"{insights['product']}<br/>{insights['product_quantity']} sold",
                        cell_style,
                    ),
                    Paragraph(f"{insights['avg_value']:,.2f}", cell_style),
                ],
            ],
            colWidths=[doc.width / 3.0] * 3,
            hAlign="LEFT",
        )
        insight_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(TEAL_PALE)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(TEAL_DARK)),
                    ("FONTNAME", (0, 0), (-1, -1), regular_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("LEADING", (0, 0), (-1, -1), 10),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(GRAY_2)),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.extend([
            summary_table,
            Spacer(1, 10),
            Paragraph("Insights", section_style),
            insight_table,
            Spacer(1, 12),
            Paragraph("Breakdown", section_style),
        ])

        if rows:
            breakdown_headers = [
                "Date",
                "Customer name",
                "Product",
                "Quantity",
                "Type",
                "Amount (PHP)",
                "Payment status",
                "Delivery status",
            ]
            breakdown_data = [[Paragraph(text, breakdown_header_style) for text in breakdown_headers]]
            for row in rows:
                breakdown_data.append(
                    [
                        Paragraph(str(row["date"]), cell_style),
                        Paragraph(str(row["customer"]), cell_style),
                        Paragraph(str(row["product"]), cell_style),
                        Paragraph(str(row["quantity"]), cell_style),
                        Paragraph(str(row["type"]), cell_style),
                        Paragraph(f"{row['amount']:,.2f}", cell_style),
                        Paragraph(str(row["payment_status"]), cell_style),
                        Paragraph(str(row["delivery_status"]), cell_style),
                    ]
                )

            breakdown_table = Table(
                breakdown_data,
                colWidths=[
                    doc.width * 0.12,
                    doc.width * 0.17,
                    doc.width * 0.13,
                    doc.width * 0.07,
                    doc.width * 0.12,
                    doc.width * 0.12,
                    doc.width * 0.14,
                    doc.width * 0.13,
                ],
                repeatRows=1,
                hAlign="LEFT",
            )
            breakdown_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(TEAL_DARK)),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1f2f2c")),
                        ("FONTNAME", (0, 0), (-1, 0), bold_font),
                        ("FONTNAME", (0, 1), (-1, -1), regular_font),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                        ("LEADING", (0, 0), (-1, -1), 9.5),
                        ("ALIGN", (3, 1), (3, -1), "CENTER"),
                        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
                        ("ALIGN", (6, 1), (7, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(GRAY_2)),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbfb")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.append(breakdown_table)
        else:
            story.append(Paragraph("No records found for the selected date range.", info_style))

        footer_text = "This report was generated by G and C LPG Trading Delivery System"
        doc.build(
            story,
            canvasmaker=lambda *args, **kwargs: NumberedCanvas(
                *args,
                footer_text=footer_text,
                footer_font_name=regular_font,
                **kwargs,
            ),
        )

    def _export_excel(self):
        default_name = f"{self.period_name.lower()}_report.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Report as Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            metadata = self._export_metadata()
            summary = self._summary_metrics()
            rows = self._report_rows()

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"

            teal_fill = XLPatternFill("solid", fgColor="145F55")
            pale_fill = XLPatternFill("solid", fgColor="E8F5F3")
            light_fill = XLPatternFill("solid", fgColor="F8FBFB")
            dark_font = XLFont(color="145F55", bold=True)
            white_bold_font = XLFont(color="FFFFFF", bold=True)
            thin_border = XLBorder(
                left=XLSide(style="thin", color="E6EAE9"),
                right=XLSide(style="thin", color="E6EAE9"),
                top=XLSide(style="thin", color="E6EAE9"),
                bottom=XLSide(style="thin", color="E6EAE9"),
            )

            sheet.merge_cells("A1:H1")
            sheet["A1"] = metadata["business_name"]
            sheet["A1"].font = XLFont(size=16, bold=True, color="145F55")
            sheet["A1"].alignment = XLAlignment(horizontal="left")

            sheet.merge_cells("A2:H2")
            sheet["A2"] = metadata["address"]
            sheet["A2"].font = XLFont(size=10, color="7A8A87")

            sheet.merge_cells("A4:H4")
            sheet["A4"] = f"{metadata['report_type']} Report"
            sheet["A4"].font = XLFont(size=14, bold=True, color="145F55")

            sheet.merge_cells("A5:H5")
            sheet["A5"] = f"Date range covered: {metadata['date_range']}"
            sheet["A5"].font = XLFont(size=10, color="3A4A47")

            sheet.merge_cells("A6:H6")
            sheet["A6"] = metadata["generated_on"]
            sheet["A6"].font = XLFont(size=10, color="3A4A47")

            sheet.merge_cells("A7:H7")
            sheet["A7"] = f"Generated by: {metadata['generated_by']}"
            sheet["A7"].font = XLFont(size=10, color="3A4A47")

            sheet["A9"] = "Summary"
            sheet["A9"].font = XLFont(size=11, bold=True, color="145F55")

            summary_headers = [
                "Total deliveries",
                "Delivered",
                "Cancelled",
                "Pending",
                "In transit",
                "Delivered sales (PHP)",
                "Total paid (PHP)",
                "Total unpaid (PHP)",
            ]
            summary_values = [
                summary["total_deliveries"],
                summary["total_delivered"],
                summary["total_cancelled"],
                summary["total_pending"],
                summary["total_in_transit"],
                f"{summary['total_sales']:,.2f}",
                f"{summary['total_paid']:,.2f}",
                f"{summary['total_unpaid']:,.2f}",
            ]
            for column_index, header_text in enumerate(summary_headers, start=1):
                cell = sheet.cell(row=10, column=column_index, value=header_text)
                cell.fill = pale_fill
                cell.font = dark_font
                cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for column_index, value in enumerate(summary_values, start=1):
                cell = sheet.cell(row=11, column=column_index, value=value)
                cell.fill = light_fill
                cell.font = XLFont(color="3A4A47")
                cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            insights = self._insight_values(summary)
            sheet["A13"] = "Insights"
            sheet["A13"].font = XLFont(size=11, bold=True, color="145F55")

            insight_headers = [
                "Peak sales day",
                "Most sold product",
                "Average order value (PHP)",
            ]
            insight_values = [
                f"{insights['peak_day']} ({insights['peak_amount']:,.2f})",
                f"{insights['product']} ({insights['product_quantity']} sold)",
                f"{insights['avg_value']:,.2f}",
            ]
            for column_index, header_text in enumerate(insight_headers, start=1):
                cell = sheet.cell(row=14, column=column_index, value=header_text)
                cell.fill = pale_fill
                cell.font = dark_font
                cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for column_index, value in enumerate(insight_values, start=1):
                cell = sheet.cell(row=15, column=column_index, value=value)
                cell.fill = light_fill
                cell.font = XLFont(color="3A4A47")
                cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            sheet["A17"] = "Breakdown"
            sheet["A17"].font = XLFont(size=11, bold=True, color="145F55")

            breakdown_headers = [
                "Date",
                "Customer name",
                "Product",
                "Quantity",
                "Type",
                "Amount (PHP)",
                "Payment status",
                "Delivery status",
            ]
            header_row = 18
            for column_index, header_text in enumerate(breakdown_headers, start=1):
                cell = sheet.cell(row=header_row, column=column_index, value=header_text)
                cell.fill = teal_fill
                cell.font = white_bold_font
                cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for row_index, row in enumerate(rows, start=header_row + 1):
                values = [
                    row["date"],
                    row["customer"],
                    row["product"],
                    row["quantity"],
                    row["type"],
                    f"{row['amount']:,.2f}",
                    row["payment_status"],
                    row["delivery_status"],
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_index, column=column_index, value=value)
                    cell.fill = light_fill if row_index % 2 == 0 else XLPatternFill(fill_type=None)
                    cell.font = XLFont(color="3A4A47")
                    cell.alignment = XLAlignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border

            widths = [14, 22, 16, 10, 12, 14, 16, 16]
            for column_index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(column_index)].width = width

            sheet.row_dimensions[10].height = 22
            sheet.row_dimensions[11].height = 22
            sheet.row_dimensions[14].height = 22
            sheet.row_dimensions[15].height = 26
            workbook.save(file_path)

            ExportNoticeDialog("Export Successful", f"Excel exported successfully to:\n{file_path}", self).exec()
        except Exception as exc:
            ExportNoticeDialog("Export Failed", f"Could not export Excel report.\n\n{exc}", self).exec()

    def _refresh_metrics(self, summary=None):
        summary = summary or self._summary_data or {}
        self._refresh_sales_labels()
        self._delivery_total_lbl.setText(str(int(summary.get("total_deliveries", 0) or 0)))
        self._delivered_value_lbl.setText(str(int(summary.get("total_delivered", 0) or 0)))
        self._pending_value_lbl.setText(str(int(summary.get("total_pending", 0) or 0)))
        self._in_transit_value_lbl.setText(str(int(summary.get("total_in_transit", 0) or 0)))
        self._cancelled_value_lbl.setText(str(int(summary.get("total_cancelled", 0) or 0)))
        self._sales_val_lbl.setText(f"PHP {float(summary.get('total_sales', 0) or 0):,.2f}")
        self._paid_value_lbl.setText(f"PHP {float(summary.get('total_paid', 0) or 0):,.2f}")
        self._unpaid_value_lbl.setText(f"PHP {float(summary.get('total_unpaid', 0) or 0):,.2f}")
        self._peak_day_value_lbl.setText(self._compact_label(summary.get("peak_sales_day"), 22))
        self._peak_day_detail_lbl.setText(f"PHP {float(summary.get('peak_sales_amount', 0) or 0):,.2f}")
        self._most_product_value_lbl.setText(self._compact_label(summary.get("most_sold_product"), 24))
        self._most_product_detail_lbl.setText(
            f"{int(summary.get('most_sold_product_quantity', 0) or 0)} sold"
        )
        self._avg_order_value_lbl.setText(f"PHP {float(summary.get('avg_transaction_value', 0) or 0):,.2f}")
        self._avg_order_detail_lbl.setText("Delivered sales per order")

    def set_report_data(self, summary, rows, from_date=None, to_date=None):
        self._summary_data = dict(summary or {})
        self._base_data = list(rows or [])
        if from_date is not None:
            self._current_from_date = from_date
        if to_date is not None:
            self._current_to_date = to_date
        self._apply_local_filters(prefer_summary=True)

class ReportView(QFrame):
    """Main Report View"""
    def __init__(self, user=None, show_topbar=True, parent=None, controller=None):
        super().__init__(parent)
        _ensure_fonts_loaded()
        self.user = user or {"full_name": "", "role": "owner"}
        self.owner_name = self.user.get("full_name") or self.user.get("username") or "Owner"
        self.show_topbar = show_topbar
        self._controller = None

        self.setStyleSheet(f"background:{GRAY_1};border:none;")
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        root_lay = QVBoxLayout(self)
        root_lay.setContentsMargins(0, 0, 0, 0)
        root_lay.setSpacing(0)

        if show_topbar:
            root_lay.addWidget(self._build_topbar())

        root_lay.addWidget(self._build_content())
        if controller:
            self.bind_controller(controller, request_initial=True)

    def _build_topbar(self):
        """Build top bar with date/time"""
        bar = QWidget()
        bar.setFixedHeight(84)
        bar.setStyleSheet(f"background:{WHITE};border:none;border-bottom:1px solid {GRAY_2};")
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(28, 8, 28, 8)

        self._breadcrumb_lbl = QLabel("REPORTS")
        self._breadcrumb_lbl.setFont(inter(11, QFont.Medium))
        self._breadcrumb_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._breadcrumb_lbl.setStyleSheet(
            f"color:{GRAY_4};letter-spacing:0.5px;background:transparent;border:none;"
        )
        lay.addWidget(self._breadcrumb_lbl)
        lay.addStretch()

        clock_col = QVBoxLayout()
        clock_col.setContentsMargins(0, 0, 0, 0)
        clock_col.setSpacing(2)
        clock_col.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self._clock_lbl = QLabel("--:--:--")
        self._clock_lbl.setFont(inter(17, QFont.Medium))
        self._clock_lbl.setMinimumHeight(30)
        self._clock_lbl.setAlignment(Qt.AlignRight)
        self._clock_lbl.setStyleSheet(
            f"color:{TEAL_DARK};letter-spacing:1px;background:transparent;border:none;"
        )

        self._date_lbl = QLabel(QDate.currentDate().toString("dddd, MMMM d, yyyy"))
        self._date_lbl.setFont(inter(11))
        self._date_lbl.setMinimumHeight(18)
        self._date_lbl.setAlignment(Qt.AlignRight)
        self._date_lbl.setStyleSheet(
            f"color:{GRAY_4};letter-spacing:0.3px;background:transparent;border:none;"
        )

        clock_col.addWidget(self._clock_lbl)
        clock_col.addWidget(self._date_lbl)
        lay.addLayout(clock_col)
        lay.addSpacing(16)

        notif = QPushButton()
        notif.setFixedSize(36, 36)
        notif.setCursor(Qt.PointingHandCursor)
        notif.setText("!")
        notif.setStyleSheet(
            f"""
            QPushButton{{
                color:{GRAY_4};background:{WHITE};
                border:1px solid {GRAY_2};border-radius:6px;font-size:14px;
            }}
            QPushButton:hover{{background:{GRAY_1};}}
        """
        )
        lay.addWidget(notif)

        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)
        self._tick()

        return bar

    def _tick(self):
        """Update clock"""
        if hasattr(self, "_clock_lbl"):
            from PySide6.QtCore import QTime
            self._clock_lbl.setText(QTime.currentTime().toString("hh:mm:ss"))

    def _configure_date_edit(self, date_edit, min_height=34, min_width=130):
        date_edit.setStyleSheet(_date_edit_style(min_height=min_height, min_width=min_width))
        calendar = date_edit.calendarWidget()
        if calendar is None:
            return

        calendar.setFont(inter(11))
        calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        calendar.setHorizontalHeaderFormat(QCalendarWidget.ShortDayNames)
        calendar.setGridVisible(False)
        calendar.setMinimumSize(320, 260)
        calendar.setStyleSheet(_calendar_native_tint_style())

        saturday_fmt = QTextCharFormat()
        saturday_fmt.setForeground(QColor("#000000"))
        calendar.setWeekdayTextFormat(Qt.Saturday, saturday_fmt)

        sunday_fmt = QTextCharFormat()
        sunday_fmt.setForeground(QColor(TEAL))
        calendar.setWeekdayTextFormat(Qt.Sunday, sunday_fmt)

        for btn_name in ("qt_calendar_monthbutton", "qt_calendar_yearbutton"):
            btn = calendar.findChild(QToolButton, btn_name)
            if btn is not None and btn.menu() is not None:
                btn.menu().setStyleSheet(_calendar_popup_menu_style())

        _attach_year_dropdown(calendar)

        prev_btn = calendar.findChild(QToolButton, "qt_calendar_prevmonth")
        if prev_btn is not None:
            prev_btn.setIcon(QIcon())
            prev_btn.setText("<")
            prev_btn.setToolButtonStyle(Qt.ToolButtonTextOnly)
            prev_btn.setStyleSheet(
                f"""
                QToolButton {{
                    color:{WHITE};
                    background:transparent;
                    border:none;
                    font-size:20px;
                    font-weight:600;
                }}
                QToolButton:hover {{
                    background:rgba(255,255,255,0.16);
                    border-radius:8px;
                }}
                """
            )

        next_btn = calendar.findChild(QToolButton, "qt_calendar_nextmonth")
        if next_btn is not None:
            next_btn.setIcon(QIcon())
            next_btn.setText(">")
            next_btn.setToolButtonStyle(Qt.ToolButtonTextOnly)
            next_btn.setStyleSheet(
                f"""
                QToolButton {{
                    color:{WHITE};
                    background:transparent;
                    border:none;
                    font-size:20px;
                    font-weight:600;
                }}
                QToolButton:hover {{
                    background:rgba(255,255,255,0.16);
                    border-radius:8px;
                }}
                """
            )

    def _build_content(self):
        """Build main content area"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet(owner_scrollbar_qss())

        w = QWidget()
        w.setStyleSheet("background:transparent;")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(28, 24, 28, 28)
        lay.setSpacing(20)

        # Header Section
        header = QHBoxLayout()
        header.setContentsMargins(0, 0, 0, 0)
        header.setSpacing(0)

        left_col = QVBoxLayout()
        left_col.setSpacing(0)

        top_lbl = QLabel("REPORTS OVERVIEW")
        top_lbl.setFont(inter(10, QFont.DemiBold))
        top_lbl.setStyleSheet(
            f"color:{TEAL};letter-spacing:2px;background:transparent;border:none;margin-bottom:5px;"
        )

        title = QLabel("Reports and Analytics")
        title.setFont(playfair(28, QFont.Medium))
        title.setStyleSheet(f"color:{TEAL_DARK};background:transparent;border:none;")

        subtitle = QLabel("Monitor delivery performance, sales totals, and report summaries by period.")
        subtitle.setFont(inter(12))
        subtitle.setWordWrap(False)
        subtitle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        subtitle.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;margin-top:4px;")

        left_col.addWidget(top_lbl)
        left_col.addWidget(title)
        left_col.addWidget(subtitle)
        header.addLayout(left_col)
        header.addStretch()
        lay.addLayout(header)

        # Tab Buttons Row
        tabs_row = QHBoxLayout()
        tabs_row.setContentsMargins(0, 0, 0, 0)
        tabs_row.setSpacing(8)

        self._tab_buttons = {}
        self._tab_pages = {}

        # Daily Tab
        today = QDate.currentDate()
        daily_btn = self._create_tab_button("   Daily   ", active=True)
        daily_btn.clicked.connect(lambda: self._switch_tab("daily"))
        self._tab_buttons["daily"] = daily_btn
        self._tab_pages["daily"] = ReportTabPanel("Daily", today, today, owner_name=self.owner_name)
        tabs_row.addWidget(daily_btn)

        # Weekly Tab
        week_start = today.addDays(-today.dayOfWeek() + 1)
        weekly_btn = self._create_tab_button("   Weekly   ")
        weekly_btn.clicked.connect(lambda: self._switch_tab("weekly"))
        self._tab_buttons["weekly"] = weekly_btn
        self._tab_pages["weekly"] = ReportTabPanel(
            "Weekly", week_start, today, owner_name=self.owner_name
        )
        tabs_row.addWidget(weekly_btn)

        # Monthly Tab
        month_start = QDate(today.year(), today.month(), 1)
        monthly_btn = self._create_tab_button("   Monthly   ")
        monthly_btn.clicked.connect(lambda: self._switch_tab("monthly"))
        self._tab_buttons["monthly"] = monthly_btn
        self._tab_pages["monthly"] = ReportTabPanel(
            "Monthly", month_start, today, owner_name=self.owner_name
        )
        tabs_row.addWidget(monthly_btn)

        tabs_row.addStretch()

        filter_frame = QFrame()
        filter_frame.setStyleSheet("background:transparent;border:none;")
        filter_frame.setFixedHeight(40)

        filter_lay = QHBoxLayout(filter_frame)
        filter_lay.setContentsMargins(12, 5, 12, 5)
        filter_lay.setSpacing(8)

        from_lbl = QLabel("From:")
        from_lbl.setFont(inter(10))
        from_lbl.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")

        self.from_date = QDateEdit()
        self.from_date.setDate(QDate.currentDate())
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("MMM d, yyyy")
        self.from_date.setFont(inter(10))
        self.from_date.setFixedWidth(124)
        self.from_date.dateChanged.connect(self._apply_filter)
        self._configure_date_edit(self.from_date, min_height=28, min_width=124)

        to_lbl = QLabel("To:")
        to_lbl.setFont(inter(9))
        to_lbl.setStyleSheet(f"color:{GRAY_4};background:transparent;border:none;")

        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("MMM d, yyyy")
        self.to_date.setFont(inter(9))
        self.to_date.setFixedWidth(124)
        self.to_date.dateChanged.connect(self._apply_filter)
        self._configure_date_edit(self.to_date, min_height=28, min_width=124)

        filter_lay.addWidget(from_lbl)
        filter_lay.addWidget(self.from_date)
        filter_lay.addWidget(to_lbl)
        filter_lay.addWidget(self.to_date)
        tabs_row.addWidget(filter_frame)
        lay.addLayout(tabs_row)

        # Tab Content Stack
        self._content_stack = QStackedWidget()
        self._content_stack.setStyleSheet("background:transparent;border:none;")
        self._content_stack.addWidget(self._tab_pages["daily"])
        self._content_stack.addWidget(self._tab_pages["weekly"])
        self._content_stack.addWidget(self._tab_pages["monthly"])
        self._content_stack.setCurrentIndex(0)
        lay.addWidget(self._content_stack)

        self._switch_tab("daily")

        lay.addStretch()
        scroll.setWidget(w)
        return scroll

    def bind_controller(self, controller, request_initial=False):
        self._controller = controller
        if hasattr(controller, "attach_view"):
            controller.attach_view(self)
        if request_initial and self._controller:
            self.reload_active_report()

    def current_date_range(self):
        def to_pydate(qdate):
            if hasattr(qdate, "toPython"):
                return qdate.toPython()
            return qdate.toPyDate()

        return to_pydate(self.from_date.date()), to_pydate(self.to_date.date())

    def current_period_name(self):
        active_tab = self._content_stack.currentWidget()
        return active_tab.period_name if active_tab is not None else "Daily"

    def reload_active_report(self):
        if not self._controller:
            return
        date_from, date_to = self.current_date_range()
        self._controller.load_period(self.current_period_name(), date_from, date_to)

    def load_report_data(self, period_name, payload, date_from, date_to):
        period_key = str(period_name or "").strip().lower()
        target_page = self._tab_pages.get(period_key)
        if target_page is None:
            return
        target_page.set_report_data(
            payload.get("summary", {}),
            payload.get("rows", []),
            QDate(date_from.year, date_from.month, date_from.day),
            QDate(date_to.year, date_to.month, date_to.day),
        )

    def show_error(self, title, message):
        QMessageBox.critical(self, title, str(message))

    def showEvent(self, event):
        super().showEvent(event)
        if self._controller:
            self.reload_active_report()

    def hideEvent(self, event):
        self.reset_view_state(reload=False)
        super().hideEvent(event)

    def reset_view_state(self, reload=True):
        self._switch_tab("daily", reload=reload)

    def _create_tab_button(self, text, active=False):
        """Create a tab button"""
        btn = QPushButton(text)
        btn.setFont(inter(10, QFont.Medium if active else QFont.Normal))
        btn.setFixedHeight(40)
        btn.setCursor(Qt.PointingHandCursor)

        if active:
            btn.setStyleSheet(
                f"""
                QPushButton{{
                    background:{TEAL};color:#fff;border:none;
                    border-radius:6px;padding:0 16px;
                }}
                QPushButton:hover{{background:{TEAL_MID};}}
            """
            )
        else:
            btn.setStyleSheet(
                f"""
                QPushButton{{
                    background:{WHITE};color:{GRAY_4};border:1px solid {GRAY_2};
                    border-radius:6px;padding:0 16px;
                }}
                QPushButton:hover{{background:{GRAY_1};}}
            """
            )

        return btn

    def _switch_tab(self, tab_name, reload=True):
        """Switch between tabs"""
        tab_index = {"daily": 0, "weekly": 1, "monthly": 2}.get(tab_name, 0)
        self._content_stack.setCurrentIndex(tab_index)

        today = QDate.currentDate()
        self.from_date.blockSignals(True)
        self.to_date.blockSignals(True)

        if tab_name == "daily":
            self.from_date.setDate(today)
            self.to_date.setDate(today)
        elif tab_name == "weekly":
            week_start = today.addDays(-(today.dayOfWeek() - 1))
            self.from_date.setDate(week_start)
            self.to_date.setDate(today)
        elif tab_name == "monthly":
            self.from_date.setDate(QDate(today.year(), today.month(), 1))
            self.to_date.setDate(today)

        self.from_date.blockSignals(False)
        self.to_date.blockSignals(False)

        if reload:
            self._apply_filter()

        # Update tab button styles
        for btn_name, btn in self._tab_buttons.items():
            if btn_name == tab_name:
                btn.setStyleSheet(
                    f"""
                    QPushButton{{
                        background:{TEAL};color:#fff;border:none;
                        border-radius:6px;padding:0 16px;
                    }}
                    QPushButton:hover{{background:{TEAL_MID};}}
                """
                )
                btn.setFont(inter(10, QFont.Medium))
            else:
                btn.setStyleSheet(
                    f"""
                    QPushButton{{
                        background:{WHITE};color:{GRAY_4};border:1px solid {GRAY_2};
                        border-radius:6px;padding:0 16px;
                    }}
                    QPushButton:hover{{background:{GRAY_1};}}
                """
                )
                btn.setFont(inter(10))

    def _apply_filter(self):
        """Apply custom date range filter to the active tab immediately."""
        from_date = self.from_date.date()
        to_date = self.to_date.date()
        if from_date > to_date:
            from_date, to_date = to_date, from_date
            self.from_date.blockSignals(True)
            self.to_date.blockSignals(True)
            self.from_date.setDate(from_date)
            self.to_date.setDate(to_date)
            self.from_date.blockSignals(False)
            self.to_date.blockSignals(False)

        active_tab = self._content_stack.currentWidget()
        if active_tab is not None and hasattr(active_tab, "apply_date_filter"):
            active_tab.apply_date_filter(from_date, to_date)
        if self._controller:
            self.reload_active_report()

    def _modern_scrollbar_qss(self):
        """Modern scrollbar styling"""
        return owner_scrollbar_qss()


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    app = QApplication([])
    view = ReportView()
    view.show()
    sys.exit(app.exec())
